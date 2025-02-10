import pymysql
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel

# 数据库连接信息
db_config = {
    "host": "localhost",
    "port": 3306,
    "database": "stock",
    "user": "root",
    "password": "123456",
}


def generate_fund_chart(fund_codes, output_path):
    """
    生成包含多基金折线图的 Excel 文件。

    :param fund_codes: 基金代码列表。
    :param output_path: 输出文件路径。
    """

    def fetch_data_from_db(fund_codes):
        """
        从数据库获取多个基金代码的累计净值数据。
        :param fund_codes: 基金代码列表。
        :return: categories, data_sets, min_value, max_value
        """
        query = """
        SELECT fund_code, net_value_date, cumulative_net_value
        FROM fund_net_value
        WHERE fund_code IN (%s)
        ORDER BY net_value_date
        """
        placeholders = ", ".join(["%s"] * len(fund_codes))
        query = query % placeholders

        connection = pymysql.connect(**db_config)
        try:
            with connection.cursor() as cursor:
                cursor.execute(query, fund_codes)
                rows = cursor.fetchall()
        finally:
            connection.close()

        # 整理数据
        data_by_code = {}
        for fund_code in fund_codes:
            data_by_code[fund_code] = {}

        categories = set()
        min_value = float("inf")
        max_value = float("-inf")

        for row in rows:
            fund_code, date, value = row
            if fund_code in data_by_code:
                if value is not None:  # 检查 value 是否为 None
                    data_by_code[fund_code][date] = float(value)  # 转换为 float
                    categories.add(date)
                    # 更新最小值和最大值
                    min_value = min(min_value, float(value))
                    max_value = max(max_value, float(value))
                else:
                    data_by_code[fund_code][date] = None  # 如果是 None，直接存储为 None

        # 将数据格式化为需要的格式
        categories = sorted(categories)
        data_sets = []
        for fund_code, values in data_by_code.items():
            data = [values.get(date, None) for date in categories]
            data_sets.append((fund_code, data))

        return categories, data_sets, min_value, max_value

    def create_line_chart(wb, data_sets, categories, title, min_value, max_value):
        """
        在 Excel 工作簿中创建折线图。

        :param wb: OpenPyXL Workbook 对象。
        :param data_sets: 数据集列表，每个元素包含标签和数据值。
        :param categories: X轴分类标签列表。
        :param title: 图表标题。
        :param min_value: 纵轴最小值。
        :param max_value: 纵轴最大值。
        """
        sheet_name = title[:30]
        counter = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f"{title[:27]}_{counter}"
            counter += 1

        # 创建新表
        sheet = wb.create_sheet(title=sheet_name)

        # 写入分类和数据集
        sheet.append(["Category"] + [label for label, _ in data_sets])
        for i, category in enumerate(categories):
            row = [category] + [data[i] for _, data in data_sets]
            sheet.append(row)

        # 创建折线图
        chart = LineChart()
        chart.title = title

        # 添加数据和分类到图表
        for col_idx, (label, _) in enumerate(data_sets, start=2):
            data_ref = Reference(sheet, min_col=col_idx, min_row=2, max_row=len(categories) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.series[-1].tx = SeriesLabel(v=label)

        categories_ref = Reference(sheet, min_col=1, min_row=2, max_row=len(categories) + 1)
        chart.set_categories(categories_ref)

        # 设置纵轴范围
        margin = (max_value - min_value) * 0.1  # 添加 10% 的上下边界
        chart.y_axis.scaling.min = max(0, min_value - margin)
        chart.y_axis.scaling.max = max_value + margin

        # 设置图表大小
        chart.width = 30
        chart.height = 20

        # 添加图表到工作表
        sheet.add_chart(chart, "K2")

    # 创建 Excel 工作簿
    wb = Workbook()
    wb.remove(wb.active)

    # 获取数据并生成图表
    categories, data_sets, min_value, max_value = fetch_data_from_db(fund_codes)
    if data_sets:
        create_line_chart(wb, data_sets, categories, "Fund Comparison", min_value, max_value)

    # 保存 Excel 文件
    wb.save(output_path)
    print(f"Chart with multiple funds created and saved to {output_path}!")


# 示例调用
if __name__ == "__main__":
    # fund_codes = ["007562"]
    fund_codes1 = ['750002', '000122', '000200', '000310', '000335', '000436', '000801', '001316']# 436(差一些) 801
    fund_codes2 = ['004010', '002920', '003978', '675111', '675113', '004063', '003657', '004839']#4 839 2920 4063(差一些) 3978(差一些) 675111(差一些) 675113(差一些) 4010(波动大，收益高)
    fund_codes3 = ['006331', '006989', '007075', '007319', '007582', '900019', '007194', '007551']# 都差不多
    fund_codes4 = ['007828', '007562', '007837', '008176', '007229', '008383', '008204', '009219']
    fund_codes = ['090010', '002611', '003376', '070009']
    output_file = "output/多折线图-无时间.xlsx"
    generate_fund_chart(fund_codes, output_file)
