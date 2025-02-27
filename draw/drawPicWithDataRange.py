import pymysql
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel

# 数据库连接信息
db_config = {
    "host": "localhost",
    "port": 3306,
    "database": "stock",
    "user": "root",
    "password": "123456",
}


def generate_fund_chart(fund_codes, output_path, date_range=None):
    """
    生成包含多基金折线图的 Excel 文件。

    :param fund_codes: 基金代码列表。
    :param output_path: 输出文件路径。
    :param date_range: 时间范围 (start_date, end_date)，格式为 YYYY-MM-DD 的字符串元组。
    """
    def fetch_data_from_db(fund_codes, date_range):
        """
        从数据库获取多个基金代码的累计净值数据。
        :param fund_codes: 基金代码列表。
        :param date_range: 时间范围 (start_date, end_date)，格式为 YYYY-MM-DD 的字符串元组。
        :return: categories, data_sets, min_value, max_value
        """
        query = """
        SELECT fund_code, net_value_date, cumulative_net_value
        FROM fund_net_value
        WHERE fund_code IN (%s)
        """
        placeholders = ", ".join(["%s"] * len(fund_codes))
        query = query % placeholders

        # 添加日期过滤条件
        params = fund_codes
        if date_range:
            query += " AND net_value_date BETWEEN %s AND %s"
            params += date_range

        query += " ORDER BY net_value_date"

        connection = pymysql.connect(**db_config)
        try:
            with connection.cursor() as cursor:
                cursor.execute(query, params)
                rows = cursor.fetchall()
        finally:
            connection.close()

        # 整理数据
        data_by_code = {}
        for fund_code in fund_codes:
            data_by_code[fund_code] = {}

        categories = set()
        min_value = float("inf")
        max_value = float("-inf")

        for row in rows:
            fund_code, date, value = row
            if fund_code in data_by_code:
                value = float(value)  # 确保转换为 float
                data_by_code[fund_code][date] = value
                categories.add(date)
                min_value = min(min_value, value)
                max_value = max(max_value, value)

        # 将数据格式化为需要的格式
        categories = sorted(categories)
        data_sets = []
        for fund_code, values in data_by_code.items():
            data = [values.get(date, None) for date in categories]
            data_sets.append((fund_code, data))

        return categories, data_sets, min_value, max_value

    def create_line_chart(wb, data_sets, categories, title, min_value, max_value):
        """
        在 Excel 工作簿中创建折线图。
        """
        sheet_name = title[:30]
        counter = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f"{title[:27]}_{counter}"
            counter += 1

        # 创建新表
        sheet = wb.create_sheet(title=sheet_name)

        # 写入分类和数据集
        sheet.append(["Category"] + [label for label, _ in data_sets])
        for i, category in enumerate(categories):
            row = [category] + [data[i] for _, data in data_sets]
            sheet.append(row)

        # 创建折线图
        chart = LineChart()
        chart.title = title
        chart.x_axis.title = "Date"
        chart.y_axis.title = "Cumulative Net Value"
        chart.x_axis.number_format = "yyyy-mm-dd"  # 设置横轴为日期格式
        chart.x_axis.majorTimeUnit = "days"  # 将横轴的单位设置为天

        # 添加数据和分类到图表
        for col_idx, (label, data) in enumerate(data_sets, start=2):
            if any(data):  # 只处理非空列
                data_ref = Reference(sheet, min_col=col_idx, min_row=2, max_row=len(categories) + 1)
                chart.add_data(data_ref, titles_from_data=True)
                chart.series[-1].tx = SeriesLabel(v=label)

        # 将日期作为分类
        categories_ref = Reference(sheet, min_col=1, min_row=2, max_row=len(categories) + 1)
        chart.set_categories(categories_ref)

        # 设置动态纵轴范围
        margin = (max_value - min_value) * 0.1  # 上下边界增加 10%
        chart.y_axis.scaling.min = max(0, min_value - margin)  # 确保最小值不低于 0
        chart.y_axis.scaling.max = max_value + margin

        # 设置图表大小
        chart.width = 30
        chart.height = 20

        # 添加图表到工作表
        sheet.add_chart(chart, "K2")

    # 创建 Excel 工作簿
    wb = Workbook()
    wb.remove(wb.active)

    # 获取数据并生成图表
    categories, data_sets, min_value, max_value = fetch_data_from_db(fund_codes, date_range)
    if data_sets:
        create_line_chart(wb, data_sets, categories, "Fund Comparison", min_value, max_value)

    # 保存 Excel 文件
    wb.save(output_path)
    print(f"Chart with multiple funds created and saved to {output_path}!")


# 示例调用
if __name__ == "__main__":
    # fund_codes = ['007551', '900019', '007194', '006989', '004063', '007075', '007828', '675111',
    #               '675113', '007837', '003978', '006331', '008383', '008204', '000801', '007562',
    #               '008176', '000310', '000335', '004010', '003657', '000436', '750002', '000200',
    #               '007229', '000122', '001316']
    fund_codes = ['090010', '002611', '003376', '070009']
    output_file = "output/多折线图-有时间.xlsx"
    date_range = ("2024-01-01", "2025-01-15")  # 指定时间范围
    generate_fund_chart(fund_codes, output_file, date_range)
