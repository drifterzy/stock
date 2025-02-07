import pymysql
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from datetime import datetime

# 数据库连接信息
db_config = {
    "host": "localhost",
    "port": 3306,
    "database": "stock",
    "user": "root",
    "password": "123456",
}

def generate_fund_chart(fund_map, output_path, date_range=None):
    """
    生成包含多基金折线图的 Excel 文件。

    :param fund_map: 基金代码和基金名称的对应关系，例如 {'007551': '基金A', '900019': '基金B'}
    :param output_path: 输出文件路径。
    :param date_range: 时间范围 (start_date, end_date)，格式为 YYYY-MM-DD 的字符串元组。
    """

    def fetch_data_from_db(fund_codes, date_range):
        """
        从数据库获取多个基金代码的累积净值数据，并进行归一化处理。
        :param fund_codes: 基金代码列表。
        :param date_range: 时间范围 (start_date, end_date)，格式为 YYYY-MM-DD 的字符串元组。
        :return: categories, data_sets, min_value, max_value
        """
        query = """
        SELECT fund_code, net_value_date, cumulative_net_value
        FROM fund_net_value
        WHERE fund_code IN (%s)
        """
        placeholders = ", ".join(["%s"] * len(fund_codes))
        query = query % placeholders

        # 添加日期过滤条件
        params = list(fund_codes)
        if date_range:
            query += " AND net_value_date BETWEEN %s AND %s"
            params += date_range

        query += " ORDER BY net_value_date"

        connection = pymysql.connect(**db_config)
        try:
            with connection.cursor() as cursor:
                cursor.execute(query, params)
                rows = cursor.fetchall()
        finally:
            connection.close()

        # 整理数据
        data_by_code = {}
        for fund_code in fund_codes:
            data_by_code[fund_code] = {}

        categories = set()
        min_value = float("inf")
        max_value = float("-inf")

        # 获取每个基金的第一个净值，作为归一化的基准
        first_values = {}

        for row in rows:
            fund_code, date, value = row
            if fund_code in data_by_code:
                value = float(value)  # 确保转换为 float
                data_by_code[fund_code][date] = value
                categories.add(date)
                min_value = min(min_value, value)
                max_value = max(max_value, value)

                # 记录每个基金的第一个净值
                if fund_code not in first_values:
                    first_values[fund_code] = value

        # 将数据格式化为需要的格式
        categories = sorted(categories)
        data_sets = []
        for fund_code, values in data_by_code.items():
            first_value = first_values[fund_code]  # 获取该基金的第一个净值
            normalized_data = [values.get(date, None) / first_value for date in categories]  # 归一化数据
            data_sets.append((fund_map[fund_code], normalized_data))

        return categories, data_sets, min_value, max_value

    def create_line_chart(wb, data_sets, categories, title, min_value, max_value):
        """
        在 Excel 工作簿中创建折线图。
        """
        sheet_name = title[:30]
        counter = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f"{title[:27]}_{counter}"
            counter += 1

        # 创建新表
        sheet = wb.create_sheet(title=sheet_name)

        # 写入分类和数据集
        sheet.append(["Category"] + [label for label, _ in data_sets])
        for i, category in enumerate(categories):
            row = [category] + [data[i] for _, data in data_sets]
            sheet.append(row)

        # 创建折线图
        chart = LineChart()
        chart.title = title
        chart.x_axis.title = "Date"
        chart.y_axis.title = "Normalized Cumulative Net Value"
        chart.x_axis.number_format = "yyyy-mm-dd"  # 设置横轴为日期格式
        chart.x_axis.majorTimeUnit = "days"  # 将横轴的单位设置为天

        # 添加数据和分类到图表
        for col_idx, (label, data) in enumerate(data_sets, start=2):
            if any(data):  # 只处理非空列
                data_ref = Reference(sheet, min_col=col_idx, min_row=2, max_row=len(categories) + 1)
                chart.add_data(data_ref, titles_from_data=True)
                chart.series[-1].tx = SeriesLabel(v=label)

        # 将日期作为分类
        categories_ref = Reference(sheet, min_col=1, min_row=2, max_row=len(categories) + 1)
        chart.set_categories(categories_ref)

        # 计算纵轴范围，避免图形集中在下部分
        min_y_value = min([min(data) for _, data in data_sets if any(data)])
        max_y_value = max([max(data) for _, data in data_sets if any(data)])

        # 设置动态纵轴范围
        margin = (max_y_value - min_y_value) * 0.2  # 上下边界增加 20%
        chart.y_axis.scaling.min = min_y_value - margin  # 设定最小值，避免图形过于集中
        chart.y_axis.scaling.max = max_y_value + margin  # 设定最大值

        # 设置图表大小
        chart.width = 30
        chart.height = 20

        # 设置所有系列的数据点样式
        for series in chart.series:
            series.marker.symbol = "circle"  # 数据点样式为圆形
            series.marker.size = 4  # 数据点大小
            series.marker.graphicalProperties.solidFill = "FFFFFF"  # 数据点填充为白色

        # 添加图表到工作表
        sheet.add_chart(chart, "K2")

    # 创建 Excel 工作簿
    wb = Workbook()
    wb.remove(wb.active)

    # 获取数据并生成图表
    fund_codes = list(fund_map.keys())
    categories, data_sets, min_value, max_value = fetch_data_from_db(fund_codes, date_range)
    if data_sets:
        create_line_chart(wb, data_sets, categories, "Fund Comparison", min_value, max_value)

    # 保存 Excel 文件
    wb.save(output_path)
    print(f"Chart with multiple funds created and saved to {output_path}!")

# 示例调用
if __name__ == "__main__":
    fund_map = {
        '090010': '红利',
        '002611': '黄金',
        '003376': '国债',
        '070009': '短债',
        '110020': '沪深300'

    }
    output_file = "./basicReport.xlsx"
    date_range = ("2024-01-01", datetime.today().strftime("%Y-%m-%d"))  # 指定时间范围
    generate_fund_chart(fund_map, output_file, date_range)




